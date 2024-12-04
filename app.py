import smtplib # permite enviar e-mails utilizando o protocolo SMTP 
from openpyxl import load_workbook
from pathlib import Path
from dotenv import load_dotenv
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os


def automatizar_planilha(caminho_arquivo='data\\Bairros.xlsx'):
    # Carregando a planilha
    wb = load_workbook(caminho_arquivo)

    # Obtendo o nome das abas
    abas = wb.sheetnames

    # selecionando a aba principal
    plan_principal = wb['Base de Dados']

    # Iterando pela aba principal e adicionando os valores nas abas correspondentes aos bairros
    for index in range(2, plan_principal.max_row):
        bairro = plan_principal.cell(row=index, column=3).value
        if not bairro in wb.sheetnames:
            wb.create_sheet(bairro)
        plan_ativa = wb[bairro]
        num_linhas = plan_ativa.max_row + 1
        for j in range(1, 4):
            plan_ativa.cell(row=num_linhas, column=j).value = plan_principal.cell(row=index, column=j).value
        plan_ativa.cell(row=num_linhas, column=1).value = plan_ativa.cell(row=num_linhas, column=1).value.strftime('%d/%m/%Y')
        

    # iterar pelas abas para adicionar as colunas
    for aba in wb:
        aba['A1'].value = 'Data de Nascimento'
        aba['B1'].value = 'Pessoa'
        aba['C1'].value = 'Bairro'

    # Salvando o arquivo
    wb.save('relatorio.xlsx')


load_dotenv('.venv\\.env')

def enviar_email():
    # Configurações do servidor SMTP
    server_smtp = 'smtp.gmail.com'
    port = 587
    sender_email = 'waldemberg.pereirac@gmail.com'
    password = os.getenv('PASSWORD')

    # configurações do email
    receive_email = 'berg.polimeros@hotmail.com'
    subject = "E-mail automático em Python"
    body = """<p>Olá!</p>
    <p>Segue o relatório de e-mail</p>"""

    # criando o e-mail
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receive_email
    message['Subject'] = subject
    message.attach(MIMEText(body, "html")) # faz a configuração do corpo do e-mail

    # Caminho do arquivo a ser anexado
    file_path = "relatorio.xlsx"  
    file_name = os.path.basename(file_path)

    # conectando o servidor smtp
    try:
        # Anexando o arquivo
        with open(file_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={file_name}",
            )
            message.attach(part)
        server = smtplib.SMTP(server_smtp, port) # configuração para conexão
        server.starttls()
        server.login(sender_email, password)
        server.sendmail(sender_email, receive_email, message.as_string())
        print('Email enviado com sucesso!')
    except Exception as e:
        print(f'Houve erro no envio: {e}')
    finally:
        server.quit()

def inicializar_sistema():
    automatizar_planilha()
    enviar_email()

inicializar_sistema()

        